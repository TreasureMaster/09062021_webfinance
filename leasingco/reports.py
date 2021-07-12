import json
from flask import (
    Blueprint, flash, g, redirect, render_template, request, url_for,
    send_from_directory, current_app
)
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill
from openpyxl.styles.colors import Color
from werkzeug.exceptions import abort

from leasingco.db import get_db

bp = Blueprint('reports', __name__)

# главная страница
@bp.route('/')
def index():
    g.user = True
    return render_template('index.html')

# Отчетности материнской компании
@bp.route('/parent')
def parent():
    delim = {1100, 1300, 1400, 1600, 1700}
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM Balance WHERE id_company=1")
    years = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
    cursor.execute("SELECT code, title FROM ReportsTitle WHERE code < 2000")
    balance, _ = get_balance(cursor.fetchall(), years)
    get_excel_reports(balance, 'parental', 'balance')

    cursor.execute("SELECT * FROM FinanceResults WHERE id_company=1")
    years = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
    cursor.execute("SELECT code, title FROM ReportsTitle WHERE code > 2000")
    finance, _ = get_finance(cursor.fetchall(), years)
    get_excel_reports(finance, 'parental', 'finance')

    return render_template('reports/report.html',
                            balance=balance, finance=finance,
                            delim=delim, post_title='материнской компании',
                            pre_title='', filename='parental')

# Отчетности дочерней компании
@bp.route('/subsidiary')
def subsidiary():
    delim = {1100, 1300, 1400, 1600, 1700}
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM Balance WHERE id_company=2")
    years = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
    cursor.execute("SELECT code, title FROM ReportsTitle WHERE code < 2000")
    balance, _ = get_balance(cursor.fetchall(), years)
    get_excel_reports(balance, 'subsidiary', 'balance')

    cursor.execute("SELECT * FROM FinanceResults WHERE id_company=2")
    years = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
    cursor.execute("SELECT code, title FROM ReportsTitle WHERE code > 2000")
    finance, _ = get_finance(cursor.fetchall(), years)
    get_excel_reports(finance, 'subsidiary', 'finance')

    return render_template('reports/report.html',
                            balance=balance, finance=finance,
                            delim=delim, post_title='дочерней компании',
                            pre_title='', filename='subsidiary')

# Консолидированные отчетности по обеим компаниям
@bp.route('/consolidation')
def consolidation():
    delim = {1100, 1300, 1400, 1600, 1700}
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * INTO #TempTable FROM Balance")
    cursor.execute("ALTER TABLE #TempTable DROP COLUMN id")
    cursor.execute("SELECT Year, SUM(c1110) AS c1110, SUM(c1120) AS c1120, SUM(c1150) AS c1150, "
                   "SUM(c1160) AS c1160, SUM(c1170) AS c1170, SUM(c1180) AS c1180, SUM(c1190) AS c1190, "
                   "SUM(c1210) AS c1210, SUM(c1220) AS c1220, SUM(c1230) AS c1230, SUM(c1240) AS c1240, "
                   "SUM(c1250) AS c1250, SUM(c1260) AS c1260, SUM(c1310) AS c1310, SUM(c1320) AS c1320, "
                   "SUM(c1340) AS c1340, SUM(c1350) AS c1350, SUM(c1360) AS c1360, SUM(c1370) AS c1370, "
                   "SUM(c1410) AS c1410, SUM(c1420) AS c1420, SUM(c1450) AS c1450, SUM(c1510) AS c1510, "
                   "SUM(c1520) AS c1520, SUM(c1530) AS c1530, SUM(c1540) AS c1540, SUM(c1550) AS c1550 "
                   "FROM #TempTable GROUP BY Year")

    years = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
    cursor.execute("DROP TABLE #TempTable")
    cursor.execute("SELECT code, title FROM ReportsTitle WHERE code < 2000")
    balance, totals_balance = get_balance(cursor.fetchall(), years, one=False)

    cursor.execute("SELECT * INTO #TempTable FROM FinanceResults")
    cursor.execute("ALTER TABLE #TempTable DROP COLUMN id")
    cursor.execute("SELECT Year, SUM(c2110) AS c2110, SUM(c2120) AS c2120, SUM(c2210) AS c2210, "
                   "SUM(c2220) AS c2220, SUM(c2310) AS c2310, SUM(c2320) AS c2320, SUM(c2330) AS c2330, "
                   "SUM(c2340) AS c2340, SUM(c2350) AS c2350, SUM(c2410) AS c2410, SUM(c2421) AS c2421, "
                   "SUM(c2430) AS c2430, SUM(c2450) AS c2450, SUM(c2460) AS c2460 "
                   "FROM #TempTable GROUP BY Year")
    years = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
    cursor.execute("DROP TABLE #TempTable")
    cursor.execute("SELECT code, title FROM ReportsTitle WHERE code > 2000")
    finance, totals_finance = get_finance(cursor.fetchall(), years, one=False)
    # Подгрузка взаимных оборотов
    cursor.execute("SELECT * INTO #TempTable FROM MutualTurnovers")
    cursor.execute("ALTER TABLE #TempTable DROP COLUMN id")
    cursor.execute("SELECT Year, SUM(c1170) AS c1170, SUM(c1230) AS c1230, SUM(c1240) AS c1240, "
                   "SUM(c1310) AS c1310, SUM(c1350) AS c1350, SUM(c1370) AS c1370, SUM(c1410) AS c1410, "
                   "SUM(c1450) AS c1450, SUM(c1510) AS c1510, SUM(c1520) AS c1520, SUM(c1550) AS c1550, "
                   "SUM(c2110) AS c2110, SUM(c2120) AS c2120, SUM(c2210) AS c2210, SUM(c2220) AS c2220, "
                   "SUM(c2310) AS c2310, SUM(c2320) AS c2320, SUM(c2330) AS c2330, SUM(c2340) AS c2340, "
                   "SUM(c2350) AS c2350, SUM(c2410) AS c2410, SUM(c2430) AS c2430, SUM(c2450) AS c2450, "
                   "SUM(c2460) AS c2460 "
                   "FROM #TempTable GROUP BY Year")
    years = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
    cursor.execute("DROP TABLE #TempTable")

    balance = get_mutual_balance(years, balance)
    finance = get_mutual_finance(years, finance)

    balance = correct_balance(balance, totals_balance)
    get_excel_reports(balance, 'consolidation', 'balance')
    get_excel_reports(finance, 'consolidation', 'finance')

    return render_template('reports/report.html',
                            balance=balance, finance=finance,
                            delim=delim, post_title='предприятий группы компаний',
                            pre_title='Консолидированный ', filename='consolidation')


@bp.route('/get_excel/<string:prefix>/<string:report>')
def get_excel(prefix=None, report=None):
    print(current_app.config['CLIENT_XLSX'])
    try:
        return send_from_directory(current_app.config['CLIENT_XLSX'], path=f'{prefix}_{report}.xlsx', as_attachment=True)
    except FileNotFoundError:
        abort(404)

# -------------------------------- extra funcs ------------------------------- #
def get_excel_reports(balance, title, report):
    wb = Workbook()
    ws = wb.active
    code = NamedStyle(name='code')
    code.font = Font(bold=True)
    code.alignment = Alignment(horizontal='center')
    wb.add_named_style(code)
    heading = [
        'Показатели',
        'Код строки',
        '2018',
        '2019',
        '2020'
    ]
    ws.append(heading)
    for c in range(1, 6):
            ws.cell(row=ws._current_row, column=c).fill = PatternFill(patternType='solid', fgColor=Color(rgb='00CCFFCC'))
            ws.cell(row=ws._current_row, column=c).font = Font(bold=True, size=13)
            ws.cell(row=ws._current_row, column=c).alignment = Alignment(horizontal='center')
    if title == 'balance':
        ws.append(['Актив', 'Код', '', '', ''])
        ws.cell(row=ws._current_row, column=2).style = code
    tab = Table(displayName='Test', ref='A1:E42')
    first_len = max(len(r['title']) for r in balance)
    print(first_len)
    lengths = [first_len] + [17]*4
    for num, col in enumerate(['A', 'B', 'C', 'D', 'E']):
        ws.column_dimensions[col].width = lengths[num]

    for raw_row in balance:
        row = [str(r) for r in raw_row.values()]
        ws.append(row if int(row[1]) >= 1000 else [row[0], '', '', '', ''])
        ws.cell(row=ws._current_row, column=2).style = code
        if raw_row['code'] in {1600, 1700, 2400}:
            for c in range(1, 6):
                ws.cell(row=ws._current_row, column=c).fill = PatternFill(patternType='solid', fgColor=Color(rgb='00FFFF99'))
            ws.cell(row=ws._current_row, column=1).alignment = Alignment(horizontal='right')
            ws.cell(row=ws._current_row, column=1).font = Font(bold=True)
        if raw_row['code'] == 1600:
            ws.append(['Пассив', 'Код', '', '', ''])
            ws.cell(row=ws._current_row, column=2).style = code
    ws.add_table(tab)
    wb.save(current_app.config['CLIENT_XLSX'] + '{}_{}.xlsx'.format(title, report))

def update_row(totals, row, idx, mark=False):
    # mark - для изменения бух.расчетов с 2020 года (включение кода 2421)
    if not mark:
        totals[idx]['2018'] += row['2018']
        totals[idx]['2019'] += row['2019']
    totals[idx]['2020'] += row['2020']

def get_balance(titles, years, one=True):
    balance = [{'title': title, 'code': code} for code, title in titles]
    totals = {}
    for year in years:
        for row in balance:
            idx = 'c' + str(row['code'])
            row.update(
                **({str(year['Year']): year[idx]} if idx in year else {str(year['Year']): 0})
            )
            if idx in {'c1100', 'c1200', 'c1600', 'c1300', 'c1400', 'c1500', 'c1700'}:
                totals[idx] = row
    if one:
        for row in balance:
            mark = row['code'] // 100 * 100
            if (1100 < row['code'] < 1600) and row['code'] != mark:
                update_row(totals, row, 'c' + str(mark))
        for row in balance:
            if row['code'] == 1100 or row['code'] == 1200:
                update_row(totals, row, 'c1600')
            if row['code'] == 1300 or row['code'] == 1400 or row['code'] == 1500:
                update_row(totals, row, 'c1700')

    return balance, totals

def correct_balance(balance, totals):
    for row in balance:
        mark = row['code'] // 100 * 100
        if (1100 < row['code'] < 1600) and row['code'] != mark:
            update_row(totals, row, 'c' + str(mark))
    for row in balance:
        if row['code'] == 1100 or row['code'] == 1200:
            update_row(totals, row, 'c1600')
        if row['code'] == 1300 or row['code'] == 1400 or row['code'] == 1500:
            update_row(totals, row, 'c1700')

    return balance


def get_finance(titles, years, one=True):
    finance = [{'title': title, 'code': code} for code, title in titles]
    totals = {}
    for year in years:
        for row in finance:
            idx = 'c' + str(row['code'])
            row.update(
                **({str(year['Year']): year[idx]} if idx in year else {str(year['Year']): 0})
            )
            if idx in {'c2100', 'c2200', 'c2300', 'c2400'}:
                totals[idx] = row
    if one:
        for row in finance:
            mark = row['code'] // 100 * 100
            if (2100 < row['code'] < 2500) and row['code'] != mark and row['code'] != 2421:
                update_row(totals, row, 'c' + str(mark))
            elif row['code'] == 2421:
                update_row(totals, row, 'c2400', mark=True)
        update_row(totals, totals['c2100'], 'c2200')
        update_row(totals, totals['c2200'], 'c2300')
        update_row(totals, totals['c2300'], 'c2400')
    return finance, totals

def correct_finance(finance, totals):
    for row in finance:
        mark = row['code'] // 100 * 100
        if (2100 < row['code'] < 2500) and row['code'] != mark and row['code'] != 2421:
            update_row(totals, row, 'c' + str(mark))
        elif row['code'] == 2421:
            update_row(totals, row, 'c2400', mark=True)
    update_row(totals, totals['c2100'], 'c2200')
    update_row(totals, totals['c2200'], 'c2300')
    update_row(totals, totals['c2300'], 'c2400')
    return finance

def get_mutual_balance(years, balance):
    for row in balance:
        for year in years:
            code = 'c' + str(row['code'])
            if code in year:
                row[str(year['Year'])] -= year[code]
    return balance

def get_mutual_finance(years, finance):
    total = {'2018': 0, '2019': 0, '2020': 0}
    for row in finance:
        if row['code'] == 2421:
            continue
        for year in years:
            code = 'c' + str(row['code'])
            if code in year:
                row[str(year['Year'])] -= year[code]
                total[str(year['Year'])] += row[str(year['Year'])]
            if not (row['code'] % 100):
                row[str(year['Year'])] = total[str(year['Year'])]
    return finance
